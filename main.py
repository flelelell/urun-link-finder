import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font
import time
import urllib.parse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

class UrunFinder:
    def __init__(self):
        self.sites = {
            'beko': 'https://www.beko.com.tr',
            'beko_anabayisi': 'https://www.bekoanabayisi.com',
            'akakce': 'https://www.akakce.com'
        }
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        self.results = {}

    def find_in_beko(self, urun_kodu, urun_adi):
        """Beko.com.tr sitesinde ürün ara"""
        try:
            # Ürün kodu ile ara
            search_url = f"{self.sites['beko']}/arama?q={urllib.parse.quote(urun_kodu)}"
            response = requests.get(search_url, headers=self.headers, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Ürün linkini bul
            product_links = soup.find_all('a', class_='product-link')
            if product_links:
                for link in product_links:
                    href = link.get('href')
                    if href:
                        if not href.startswith('http'):
                            href = self.sites['beko'] + href
                        return href
            
            # Ürün adı ile ara
            search_url = f"{self.sites['beko']}/arama?q={urllib.parse.quote(urun_adi)}"
            response = requests.get(search_url, headers=self.headers, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            product_links = soup.find_all('a', class_='product-link')
            if product_links:
                for link in product_links:
                    href = link.get('href')
                    if href:
                        if not href.startswith('http'):
                            href = self.sites['beko'] + href
                        return href
            
            return "Bulunamadı"
        except Exception as e:
            print(f"Beko araması hata: {e}")
            return "Bulunamadı"

    def find_in_beko_anabayisi(self, urun_kodu, urun_adi):
        """Bekoanabayisi.com sitesinde ürün ara"""
        try:
            # Ürün kodu ile ara
            search_url = f"{self.sites['beko_anabayisi']}/arama?q={urllib.parse.quote(urun_kodu)}"
            response = requests.get(search_url, headers=self.headers, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Ürün linkini bul
            product_items = soup.find_all('div', class_='product-item')
            if product_items:
                for item in product_items:
                    link = item.find('a')
                    if link and link.get('href'):
                        href = link.get('href')
                        if not href.startswith('http'):
                            href = self.sites['beko_anabayisi'] + href
                        return href
            
            # Ürün adı ile ara
            search_url = f"{self.sites['beko_anabayisi']}/arama?q={urllib.parse.quote(urun_adi)}"
            response = requests.get(search_url, headers=self.headers, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            product_items = soup.find_all('div', class_='product-item')
            if product_items:
                for item in product_items:
                    link = item.find('a')
                    if link and link.get('href'):
                        href = link.get('href')
                        if not href.startswith('http'):
                            href = self.sites['beko_anabayisi'] + href
                        return href
            
            return "Bulunamadı"
        except Exception as e:
            print(f"Beko Anabayisi araması hata: {e}")
            return "Bulunamadı"

    def find_in_akakce(self, urun_kodu, urun_adi):
        """Akakce.com sitesinde ürün ara"""
        try:
            # Ürün kodu ile ara
            search_url = f"{self.sites['akakce']}/arama/?q={urllib.parse.quote(urun_kodu)}"
            response = requests.get(search_url, headers=self.headers, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Ürün linkini bul
            product_links = soup.find_all('a', class_='productName')
            if product_links:
                for link in product_links:
                    href = link.get('href')
                    if href:
                        if not href.startswith('http'):
                            href = self.sites['akakce'] + href
                        return href
            
            # Ürün adı ile ara
            search_url = f"{self.sites['akakce']}/arama/?q={urllib.parse.quote(urun_adi)}"
            response = requests.get(search_url, headers=self.headers, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            product_links = soup.find_all('a', class_='productName')
            if product_links:
                for link in product_links:
                    href = link.get('href')
                    if href:
                        if not href.startswith('http'):
                            href = self.sites['akakce'] + href
                        return href
            
            return "Bulunamadı"
        except Exception as e:
            print(f"Akakce araması hata: {e}")
            return "Bulunamadı"

    def process_excel(self, input_file, output_file):
        """Excel dosyasını oku ve işle"""
        try:
            # Excel dosyasını aç
            wb = openpyxl.load_workbook(input_file)
            ws = wb.active
            
            # Yeni sütunlar ekle
            ws['C1'] = 'Beko Link'
            ws['D1'] = 'Beko Anabayisi Link'
            ws['E1'] = 'Akakce Link'
            
            # Başlık satırını kalınlaştır
            for cell in ['C1', 'D1', 'E1']:
                ws[cell].font = Font(bold=True)
            
            # Her satırı işle
            for row in range(2, ws.max_row + 1):
                urun_kodu = ws[f'A{row}'].value
                urun_adi = ws[f'B{row}'].value
                
                if not urun_kodu and not urun_adi:
                    continue
                
                print(f"İşleniyor: Kod={urun_kodu}, Adı={urun_adi}")
                
                # Her siteyi ara
                beko_link = self.find_in_beko(str(urun_kodu) if urun_kodu else '', str(urun_adi) if urun_adi else '')
                time.sleep(1)  # Rate limiting
                
                beko_anabayisi_link = self.find_in_beko_anabayisi(str(urun_kodu) if urun_kodu else '', str(urun_adi) if urun_adi else '')
                time.sleep(1)  # Rate limiting
                
                akakce_link = self.find_in_akakce(str(urun_kodu) if urun_kodu else '', str(urun_adi) if urun_adi else '')
                time.sleep(1)  # Rate limiting
                
                # Sonuçları Excel'e yaz
                ws[f'C{row}'] = beko_link
                ws[f'D{row}'] = beko_anabayisi_link
                ws[f'E{row}'] = akakce_link
                
                print(f"  Beko: {beko_link}")
                print(f"  Beko Anabayisi: {beko_anabayisi_link}")
                print(f"  Akakce: {akakce_link}")
            
            # Dosyayı kaydet
            wb.save(output_file)
            print(f"\n✅ Sonuçlar kaydedildi: {output_file}")
            
        except Exception as e:
            print(f"❌ Excel işleme hatası: {e}")

if __name__ == "__main__":
    finder = UrunFinder()
    
    # Giriş ve çıkış dosyaları
    input_file = "input.xlsx"  # A: Ürün Kodu, B: Ürün Adı
    output_file = "output.xlsx"
    
    print("🔍 Ürün Arama Başladı...\n")
    finder.process_excel(input_file, output_file)
    print("\n✅ İşlem Tamamlandı!")
